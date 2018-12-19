#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: set fileencoding=latin1:

import xlrd
import argparse, random
from xlutils.copy import copy
import openpyxl
from collections import OrderedDict
from pyexcel_xls import save_data
from pyexcel_xls import get_data


def new_sheet(file, workbook, all_sheet):
    # 打开文jian
    # copy文件
    all_sheet = workbook.sheet_names()

    newb = copy(workbook)
    sheet_page = 1
    try:
        if sheet_page == 1:
            sheet_page += 1
            l = workbook.sheet_names()
            for m in l:
                sheet = workbook.sheet_by_name(m)
                first_row = sheet.row_values(0)
                # 新建各个机型的sheet表
                for i in first_row[19:33]:
                    sheetnow = newb.add_sheet(i)
                    newb.save(file)
        # print len(all_sheet)
        # print all_sheet

    except Exception as e:
        # print e
        pass
    # sheet_context(file,workbook, all_sheet)


def sheet_context(file, workbook, all_sheet):
    workbook = xlrd.open_workbook(file)
    all_sheet = workbook.sheet_names()

    newb = copy(workbook)
    # 取第一张数据表
    data_sheet = workbook.sheet_by_name(all_sheet[0])
    # 获得数据表中lv列的值
    col_lv = data_sheet.col_values(3)
    # 获得每个机型列的值
    sheet_num=1
    # print len(all_sheet)
    for moto_type in range(19,33):
        # print sheet_num
        every_sheet = newb.get_sheet(all_sheet[sheet_num])

        everymoto_cols = data_sheet.col_values(moto_type)
        k=0
        len_lv = len(col_lv)
        list_sec = []
        while True:
            # 取出lv列每个值的下标
            for i in range(k, len_lv):
                if col_lv[i] == '1'and (len(everymoto_cols[i])!=0):
                    a = i
                    for j in range(i, len_lv):
                        if col_lv[j] == "1" and (len(everymoto_cols[j])==0):
                            k =j+1
                            list_sec.extend([sec for sec in range(a,j)])

                            break
                    break
            if i >= len_lv-1 or j >= len_lv -2:
                break
        # 14张表应从第一张表索取的数据行下标
        # print list_sec
        new_rows = {}
        num = 1
        for row in list_sec:
            rows_1 = data_sheet.row_values(row)
            # new_rows[num]=rows_1
            new_rows[num] = rows_1[:19]
            num += 1

        nums = [g for g in new_rows]
        lnew_rows = []
        nums.sort()
        for x in nums:
            h = [int(x)]
            for lsd in new_rows[x]:
                h.append(lsd)
            lnew_rows.append(h)
        for vc, vb in enumerate(lnew_rows):
            for vf, vg in enumerate(vb):
                every_sheet.write(vc, vf, vg)
        sheet_num += 1
        if sheet_num > 14:
            break

    newb.save(file)
    # insert_nummer(file, workbook, all_sheet)


def insert_nummer(file, workbook, all_sheet):
    workbook = xlrd.open_workbook(file)
    all_sheet = workbook.sheet_names()

    newb = copy(workbook)
    data_sheet = workbook.sheet_by_name(all_sheet[0])

    wnum = xlrd.open_workbook("num.xls")
    sheet_names = wnum.sheet_names()
    sec_name = {}
    for sheet_name in sheet_names:
        sheetnu = wnum.sheet_by_name(sheet_name)
        # 获取第一列sec号
        cols_1 = sheetnu.col_values(0)
        cols_2 = sheetnu.col_values(1)
    for i in range(0, 90):
        sec_name[cols_1[i]] = cols_2[i]
    # print sec_name.keys()
    # 获得机车型号
    motor_type14 = data_sheet.row_values(0)[19:]
    for she_index in range(1, 15):
        oldwbs = workbook.sheet_by_index(she_index)

        get_sheet = newb.get_sheet(she_index)
        all_sheets = workbook.sheet_by_name(all_sheet[she_index])
        # 获得各个表的F**
        sheet_f = all_sheets.col_values(3)
        val_f = list(set(sheet_f))
        val_f.sort()
        # print val_f
        a = val_f[-1]
        b = val_f[-2]
        val_f[-1] = b
        val_f[-2] = a
        # val_f = val_f[1:]
        ff_index = []
        for vc in val_f:
            ff_index.append(sheet_f.index(vc))
        for start_posi in ff_index:
            # print sheet_f[start_posi]
            for col_posi in range(0, 35):
                if col_posi == 4:
                    get_sheet.write(start_posi, 4, '0')
                elif col_posi == 6:
                    aa = random.randint(1000, 9999)
                    # get_sheet.write(start_posi, 6, 'F'+str(aa))
                    get_sheet.write(start_posi, 6, sheet_f[start_posi] + '_' + str(aa))
                elif col_posi == 10:
                    t = sec_name[sheet_f[start_posi]]
                    # print t
                    get_sheet.write(start_posi, 10, t)

                else:
                    get_sheet.write(start_posi, col_posi, " ")

            for rowIndex in range(start_posi, oldwbs.nrows):
                for colIndex in range(oldwbs.ncols):
                    get_sheet.write(rowIndex + 1, colIndex, oldwbs.cell(rowIndex, colIndex).value)
                    # print '*****'
        newb.save(file)
        # print "========================="
    # if_equal(file, workbook, all_sheet)


def if_equal(file, workbook, all_sheet):
    workbook = xlrd.open_workbook(file)
    all_sheet = workbook.sheet_names()

    newb = copy(workbook)
    data_sheet = workbook.sheet_by_index(0)
    motor_type14 = data_sheet.row_values(0)[19:]
    # 获得原生表F**列的值
    sum_f = data_sheet.col_values(2)
    # 去重之后所有F**值
    f_val = list(set(sum_f))
    # print len(f_val)
    f_val.sort()
    f_val = f_val[:-1]
    a = f_val[-1]
    b= f_val[-2]
    f_val[-1]=b
    f_val[-2]=a
    # print len(f_val)
    # 记录原生表中各个F**的下标
    f_index = []
    for vn in f_val:
        f_index.append(sum_f.index(vn))
    # print f_index
    motor_dict = {}
    len_row = len(data_sheet.row_values(0))
    for motor in range(19,len_row):
        motor_value=data_sheet.col_values(motor)
        motor_dict[motor_value[0]] = motor_value[1:]
    # 不同型号在不同段的值【【机型1在F00的值】【机型2在F00的值】【】。。。【】【】【】】
    motor_sums = []
    for xx in range(0, len(f_index)):
        if xx ==len(f_index)-1:
            font = f_index[xx]
            back = len(sum_f)
        else:
            font = f_index[xx]
            back = f_index[xx+1]
        for jj in motor_type14:
            motor_sums.append(motor_dict[jj][font:back])
    bf = 0
    df = len(motor_type14)
    # 每个F**的14个车型值
    # [[[机1f0][机2f0]] [[]] [[]]]
    list12 = []
    while bf < len(motor_sums):
        list34 =[]
        list34 = motor_sums[bf:df]
        list12.append(list34)
        bf, df =df, df+14
        if df >= len(motor_sums):
            df = len(motor_sums)
    # print len(list12)
    for rr in list12:
        for hjk in range(1, len(rr)):
            for bk in range(0, hjk):
                if rr[bk] == rr[hjk]:
                    # print bk, '***', hjk, '&&&', list12.index(rr)  # 第bk车型与第hjk车型在F**上相同
                    # print f_val[list12.index(rr)]
                    hjk_sheet = newb.get_sheet(hjk+1)
                    hjksheet = workbook.sheet_by_name(all_sheet[hjk+1])
                    hjk_value = hjksheet.col_values(6)
                    hjk_f = hjksheet.col_values(3)
                    if f_val[list12.index(rr)] not in hjk_f:
                        break
                    hjk_point = hjk_f.index(f_val[list12.index(rr)])
                    change_value = hjk_value[hjk_point -1]
                    # print change_value,"*&*&*&***&"
                    bk_sheet = newb.get_sheet(bk+1)
                    bksheet = workbook.sheet_by_name(all_sheet[bk+1])
                    bk_value = bksheet.col_values(6)
                    bk_f = bksheet.col_values(3)
                    bk_point = bk_f.index(f_val[list12.index(rr)])
                    correct = bk_value[bk_point-1]
                    hjk_value[hjk_point - 1] = correct
                    # print hjk_point-1
                    hjk_sheet.write(hjk_point-1, 6, correct)
                    # print hjk_value[hjk_point -1],"^%^%^^%"
                    newb.save(file)

                    break
    # first_row(file, workbook,all_sheet)


def first_row(file, workbook, all_sheet):

    workbook = xlrd.open_workbook(file)
    all_sheet = workbook.sheet_names()

    sheet_first = workbook.sheet_by_name(all_sheet[0])
    # 获取第一行值
    newWb = copy(workbook)
    rows_1 = sheet_first.row_values(0)
    # print len(rows_1)

    lens = len(all_sheet)
    for j in range(1,lens):
        oldWbs = workbook.sheet_by_index(j)
        newWs = newWb.get_sheet(j)
        for i in range(len(rows_1)-14):
            inserRowNo = 0
            newWs.write(inserRowNo, i+1, rows_1[i])
        for rowIndex in range(inserRowNo, oldWbs.nrows):
            for colIndex in range(oldWbs.ncols):
                newWs.write(rowIndex + 1, colIndex, oldWbs.cell(rowIndex, colIndex).value)
        newWb.save(file)


def read_xls_file(file):
    data = get_data(file)
    new_date = OrderedDict()
    # sheet表的数据
    for sheet_n in data.keys():
        # print(sheet_n)  # 表名
        # # 添加sheet表
        new_date.update({"Sheet1": data[sheet_n]})
        # 保存成xls文件
        save_data("%s.xls" % sheet_n, new_date)


# def devidecel(file):
#     workbook = xlrd.open_workbook("test.xlsx")
#     #
#     all_sheet = workbook.sheet_names()  # 获取所有表名
#     #     新建excel
#     for i in range(0, len(all_sheet)):
#         wb2 = openpyxl.Workbook()
#         wb2.save(all_sheet[i] + '.xlsx')
#         #         读取数据
#         wb1 = openpyxl.load_workbook(file)
#         wb2 = openpyxl.load_workbook(all_sheet[i] + '.xlsx')
#         # 获取sheet页
#         sheets1 = wb1.get_sheet_names()
#         sheets2 = wb2.get_sheet_names()
#         sheet1 = wb1.get_sheet_by_name(sheets1[i])
#         sheet2 = wb2.get_sheet_by_name(sheets2[0])
#         # 最大行数
#         max_row = sheet1.max_row
#         max_column = sheet1.max_column
#
#         for m in range(1, max_row + 1):
#             for n in range(97, 97 + max_column):
#                 n = chr(n)
#                 j = '%s%d' % (n, m)
#                 cell1 = sheet1[j].value
#                 sheet2[j].value = cell1
#             wb2.save(all_sheet[i] + '.xlsx')
#             wb1.close()
#             wb2.close()

def maines(file):
    workbook = xlrd.open_workbook(file)
    newb = copy(workbook)
    all_sheet = workbook.sheet_names()

    # 1.新建各个表
    new_sheet(file, workbook, all_sheet)
    # 2.插入每个表数据
    sheet_context(file, workbook, all_sheet)
    #3.比较F**插入编号，但未取相同值
    insert_nummer(file, workbook, all_sheet)
    #4.数据插入完毕，比较sec段值，取相同编号
    if_equal(file, workbook, all_sheet)
    #5.加入第一行
    first_row(file, workbook, all_sheet)
    # 6.拆分sheet
    read_xls_file(file)
    # print 'success'
    pass


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('file')
    args = parser.parse_args()
    maines(args.file)